#!/usr/bin/env python3
"""
Analyze Excel file and generate statistics report.
"""

import json
import sys
from pathlib import Path
from datetime import datetime
from collections import Counter

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def analyze_column(values: list) -> dict:
    """Analyze a single column of data."""
    # Filter out empty values
    non_empty = [v for v in values if v is not None and v != ""]

    if not non_empty:
        return {"type": "empty", "count": 0}

    # Determine column type
    numeric_values = []
    text_values = []

    for v in non_empty:
        if isinstance(v, (int, float)):
            numeric_values.append(v)
        else:
            text_values.append(str(v))

    if len(numeric_values) > len(text_values):
        # Numeric column
        return {
            "type": "numeric",
            "count": len(numeric_values),
            "sum": sum(numeric_values),
            "avg": sum(numeric_values) / len(numeric_values),
            "min": min(numeric_values),
            "max": max(numeric_values),
        }
    else:
        # Text column
        value_counts = Counter(text_values)
        return {
            "type": "text",
            "count": len(text_values),
            "unique": len(value_counts),
            "top_values": dict(value_counts.most_common(5)),
        }


def main():
    """Main entry point."""
    stdin_content = sys.stdin.read().strip()

    try:
        input_data = json.loads(stdin_content)
        file_path = input_data.get("file_path", "")
    except json.JSONDecodeError:
        file_path = stdin_content

    if not file_path:
        print(json.dumps({"error": "No file path provided"}, ensure_ascii=False))
        return

    path = Path(file_path).expanduser()

    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}, ensure_ascii=False))
        return

    if not HAS_OPENPYXL:
        print(json.dumps({
            "error": "openpyxl not installed. Install with: pip install openpyxl"
        }, ensure_ascii=False))
        return

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        analysis = {
            "file_info": {
                "path": str(path),
                "name": path.name,
                "sheets": len(wb.sheetnames),
            },
            "sheets_analysis": []
        }

        total_rows = 0
        total_cols = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            total_rows += max_row
            total_cols = max(total_cols, max_col)

            # Read all data for analysis
            all_data = []
            headers = []

            for i, row in enumerate(ws.iter_rows(max_row=min(max_row, 1000)), 1):
                row_data = [cell.value for cell in row]
                if i == 1:
                    headers = [str(h) if h else f"Col{j}" for j, h in enumerate(row_data, 1)]
                all_data.append(row_data)

            # Analyze each column
            columns_analysis = []
            if all_data and len(all_data) > 1:
                for col_idx in range(len(headers)):
                    col_values = [row[col_idx] for row in all_data[1:] if col_idx < len(row)]
                    col_analysis = analyze_column(col_values)
                    col_analysis["header"] = headers[col_idx] if col_idx < len(headers) else f"Col{col_idx+1}"
                    columns_analysis.append(col_analysis)

            analysis["sheets_analysis"].append({
                "name": sheet_name,
                "rows": max_row,
                "cols": max_col,
                "headers": headers,
                "columns": columns_analysis,
            })

        wb.close()

        analysis["summary"] = {
            "total_sheets": len(wb.sheetnames),
            "total_rows": total_rows,
            "max_columns": total_cols,
            "analyzed_at": datetime.now().isoformat(),
        }

        print(json.dumps(analysis, ensure_ascii=False, indent=2))

    except Exception as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
