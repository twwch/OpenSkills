#!/usr/bin/env python3
"""
Read Excel file content.

Extracts data from xlsx files including multiple sheets.
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def main():
    """Main entry point."""
    stdin_content = sys.stdin.read().strip()

    try:
        input_data = json.loads(stdin_content)
        file_path = input_data.get("file_path", "")
        sheet_name = input_data.get("sheet_name", None)
    except json.JSONDecodeError:
        file_path = stdin_content
        sheet_name = None

    if not file_path:
        print(json.dumps({"error": "No file path provided"}, ensure_ascii=False))
        return

    path = Path(file_path).expanduser()

    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}, ensure_ascii=False))
        return

    if not HAS_OPENPYXL:
        print(json.dumps({
            "status": "warning",
            "message": "openpyxl not installed. Install with: pip install openpyxl",
            "file_path": str(path),
            "file_size": path.stat().st_size,
        }, ensure_ascii=False, indent=2))
        return

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        sheets_data = []
        target_sheets = [sheet_name] if sheet_name else wb.sheetnames

        for name in target_sheets:
            if name not in wb.sheetnames:
                continue

            ws = wb[name]

            # Get dimensions
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0

            # Read data (limit rows for output size)
            rows = []
            headers = []

            for i, row in enumerate(ws.iter_rows(max_row=min(max_row, 100)), 1):
                row_data = []
                for cell in row:
                    value = cell.value
                    if value is None:
                        row_data.append("")
                    elif isinstance(value, (int, float)):
                        row_data.append(value)
                    else:
                        row_data.append(str(value))

                if i == 1:
                    headers = row_data
                rows.append(row_data)

            sheets_data.append({
                "name": name,
                "rows": max_row,
                "cols": max_col,
                "headers": headers,
                "data": rows[:50],  # Limit data for output
                "preview_rows": min(50, len(rows)),
            })

        wb.close()

        result = {
            "status": "success",
            "file_path": str(path),
            "file_name": path.name,
            "sheet_count": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
            "sheets": sheets_data,
        }

        print(json.dumps(result, ensure_ascii=False, indent=2))

    except Exception as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
